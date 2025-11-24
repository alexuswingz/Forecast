import pandas as pd
from pathlib import Path

# Read the CSV
csv_path = Path("hydrangea_weekly_metrics_with_conversion.csv")
if not csv_path.exists():
    print("CSV file not found. Run build_hydrangea_metrics.py first.")
    exit(1)

df = pd.read_csv(csv_path)

print(f"Loaded {len(df)} rows with {len(df.columns)} columns")
print(f"\nColumns: {list(df.columns)}")

# Create definitions dataframe
definitions_data = {
    'Metric': [
        'Week Start',
        'Units',
        'Sales',
        'Sessions',
        'Organic Conversion Rate',
        'Ad Impressions',
        'Ad Clicks',
        'Ad Spend',
        'Ad Sales',
        'Ad Orders',
        'Ad Units',
        'Ad Conversion Rate',
        'Total Inventory',
        'Available Inventory',
        'Reserved Inventory',
        'Inbound Working',
        'Inbound Shipped',
        'Inbound Receiving',
        'Research Inventory',
        'FBA Inventory',
        'AWD Inventory',
        'Ad CPC',
        'TACOS',
        'Organic Sales %',
    ],
    'Formula/Calculation': [
        'Start date of the week (Monday)',
        'Total units sold',
        'Total sales revenue',
        'Total page sessions (traffic)',
        '(Units Ordered / Sessions) x 100',
        'Total ad impressions',
        'Total ad clicks',
        'Total ad spend',
        '7-day attributed ad sales',
        '7-day attributed ad orders',
        '7-day attributed ad units',
        '(Ad Orders / Ad Clicks) x 100',
        'Total units across all fulfillment types',
        'Units ready to ship',
        'Units reserved for orders',
        'Units in inbound working status',
        'Units shipped to fulfillment centers',
        'Units being received at fulfillment centers',
        'Units under research/investigation',
        'Total units in FBA fulfillment centers',
        'Total units in Amazon Warehousing & Distribution',
        'Ad Spend / Ad Clicks',
        '(Ad Spend / Total Sales) x 100',
        '((Total Sales - Ad Sales) / Total Sales) x 100',
    ],
    'Source': [
        'Calculated',
        'order_items table',
        'order_items table',
        'child_traffic_metrics table',
        'child_traffic_metrics table',
        'ad_product_performance table',
        'ad_product_performance table',
        'ad_product_performance table',
        'ad_product_performance table',
        'ad_product_performance table',
        'ad_product_performance table',
        'ad_product_performance table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table',
        'inventory_snapshots table (filtered)',
        'inventory_snapshots table (filtered)',
        'Calculated',
        'Calculated',
        'Calculated',
    ],
    'Format': [
        'Date',
        'Integer',
        'Currency ($)',
        'Integer',
        'Percentage (%)',
        'Integer',
        'Integer',
        'Currency ($)',
        'Currency ($)',
        'Integer',
        'Integer',
        'Percentage (%)',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Integer',
        'Currency ($)',
        'Percentage (%)',
        'Percentage (%)',
    ]
}
definitions_df = pd.DataFrame(definitions_data)

# Create Excel with formatting
excel_path = Path("Hydrangea_Weekly_Metrics.xlsx")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write definitions sheet first
    definitions_df.to_excel(writer, sheet_name='Definitions', index=False)
    
    # Write weekly metrics sheet
    df.to_excel(writer, sheet_name='Weekly Metrics', index=False)
    
    # Get the workbook and worksheets
    workbook = writer.book
    definitions_ws = writer.sheets['Definitions']
    worksheet = writer.sheets['Weekly Metrics']
    
    # Format headers (bold)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Format Definitions sheet headers
    for cell in definitions_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format Weekly Metrics sheet headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format number columns
    from openpyxl.styles import numbers
    
    # Define column formats
    currency_cols = ['sales', 'ad_spend', 'ad_sales']
    percent_cols = ['organic_conversion_rate', 'ad_conversion_rate', 'tacos', 'organic_sales_pct']
    decimal_cols = ['ad_cpc']
    integer_cols = [
        'units', 'sessions', 'ad_impressions', 'ad_clicks', 'ad_orders', 'ad_units',
        'total_inventory', 'available_inventory', 'reserved_inventory',
        'inbound_working', 'inbound_shipped', 'inbound_receiving',
        'research_inventory', 'fba_inventory', 'awd_inventory'
    ]
    
    # Apply formats
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = worksheet.cell(1, col_idx).column_letter
        
        if col_name in currency_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '"$"#,##0.00'
        elif col_name in percent_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '0.00"%"'
        elif col_name in decimal_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '0.00'
        elif col_name in integer_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '#,##0'
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    worksheet.freeze_panes = 'A2'
    
    # Format Definitions sheet - adjust column widths
    definitions_ws.column_dimensions['A'].width = 25  # Metric
    definitions_ws.column_dimensions['B'].width = 50  # Formula/Calculation
    definitions_ws.column_dimensions['C'].width = 35  # Source
    definitions_ws.column_dimensions['D'].width = 15  # Format
    
    # Wrap text in definitions sheet
    for row in definitions_ws.iter_rows(min_row=2, max_row=len(definitions_df) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Freeze first row in definitions sheet
    definitions_ws.freeze_panes = 'A2'

print(f"\n[SUCCESS] Excel file created: {excel_path}")
print(f"\nColumns included:")
for col in df.columns:
    print(f"  - {col}")

# Verify conversion rate columns are present
if 'organic_conversion_rate' in df.columns:
    non_null = df['organic_conversion_rate'].notna().sum()
    print(f"\n[OK] Organic Conversion Rate: {non_null} weeks with data")
else:
    print("\n[ERROR] Organic Conversion Rate column missing!")

if 'ad_conversion_rate' in df.columns:
    non_null = df['ad_conversion_rate'].notna().sum()
    print(f"[OK] Ad Conversion Rate: {non_null} weeks with data")
else:
    print("[ERROR] Ad Conversion Rate column missing!")

