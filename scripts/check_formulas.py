"""
Extract and compare formulas from Excel v1.1 vs v1.2
"""
import openpyxl
from openpyxl.utils import get_column_letter

def extract_formulas():
    """Extract formulas from both versions"""
    
    print("="*100)
    print("EXCEL FORMULA EXTRACTION AND COMPARISON")
    print("="*100)
    
    # Load workbooks
    wb11 = openpyxl.load_workbook("1000 Bananas AUTOFORECAST V1.1.xlsx")
    wb12 = openpyxl.load_workbook("1000 Bananas AUTOFORECAST V1.2.xlsx")
    
    sheet11 = wb11['Forecast']
    sheet12 = wb12['Forecast']
    
    print(f"\nv1.1 Forecast: {sheet11.max_row} rows x {sheet11.max_column} cols")
    print(f"v1.2 Forecast: {sheet12.max_row} rows x {sheet12.max_column} cols")
    
    # Find header row and data start
    print("\n" + "="*100)
    print("IDENTIFYING STRUCTURE")
    print("="*100)
    
    print("\nFirst 5 rows of column A (v1.2):")
    for row in range(1, 6):
        val = sheet12[f"A{row}"].value
        print(f"  Row {row}: {val}")
    
    # Get all column headers (row 2 seems to be headers)
    print("\n" + "="*100)
    print("COLUMN HEADERS (Row 2)")
    print("="*100)
    
    headers_v12 = {}
    for col in range(1, 20):  # Check first 20 columns
        col_letter = get_column_letter(col)
        header = sheet12[f"{col_letter}2"].value
        if header:
            headers_v12[col_letter] = header
            print(f"  {col_letter}: {header}")
    
    # Extract formulas from row 3 (first data row)
    print("\n" + "="*100)
    print("FORMULAS IN ROW 3 (v1.1)")
    print("="*100)
    
    formulas_v11 = {}
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        cell = sheet11[f"{col_letter}3"]
        header = sheet11[f"{col_letter}2"].value
        
        # Check for formula
        formula = None
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
        
        if formula or header:
            formulas_v11[col_letter] = formula
            print(f"\n{col_letter}: {header}")
            if formula:
                print(f"  Formula: {formula}")
            else:
                print(f"  Value: {cell.value}")
    
    print("\n" + "="*100)
    print("FORMULAS IN ROW 3 (v1.2)")
    print("="*100)
    
    formulas_v12 = {}
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        cell = sheet12[f"{col_letter}3"]
        header = sheet12[f"{col_letter}2"].value
        
        # Check for formula
        formula = None
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
        
        if formula or header:
            formulas_v12[col_letter] = formula
            print(f"\n{col_letter}: {header}")
            if formula:
                print(f"  Formula: {formula}")
            else:
                print(f"  Value: {cell.value}")
    
    # Compare
    print("\n" + "="*100)
    print("FORMULA CHANGES DETECTED")
    print("="*100)
    
    all_cols = set(formulas_v11.keys()) | set(formulas_v12.keys())
    changes_found = False
    
    for col_letter in sorted(all_cols):
        v11_formula = formulas_v11.get(col_letter)
        v12_formula = formulas_v12.get(col_letter)
        
        if v11_formula != v12_formula:
            changes_found = True
            header_v11 = sheet11[f"{col_letter}2"].value
            header_v12 = sheet12[f"{col_letter}2"].value
            
            print(f"\n[CHANGE] Column {col_letter}: {header_v12 or header_v11}")
            print(f"  v1.1: {v11_formula or 'No formula'}")
            print(f"  v1.2: {v12_formula or 'No formula'}")
    
    if not changes_found:
        print("\n[NO CHANGES] All formulas are identical between v1.1 and v1.2")
    
    # Check specific critical columns
    print("\n" + "="*100)
    print("CRITICAL FORECAST COLUMNS VERIFICATION")
    print("="*100)
    
    critical_checks = [
        ('D', 'units_smooth_env'),
        ('E', 'units_final_smooth_env'),
        ('F', 'units_final_smooth'),
        ('G', 'forecast'),
        ('H', 'forecast_units_peak_env'),
        ('I', 'forecast_final_smooth'),
        ('J', 'sales_velocity_adj_weighted')
    ]
    
    for col_letter, expected_name in critical_checks:
        header_v12 = sheet12[f"{col_letter}2"].value
        cell_v12 = sheet12[f"{col_letter}3"]
        
        formula = None
        if hasattr(cell_v12, 'value') and isinstance(cell_v12.value, str) and cell_v12.value.startswith('='):
            formula = cell_v12.value
        
        print(f"\n[{col_letter}] {header_v12} (expected: {expected_name}):")
        if formula:
            print(f"  Formula (v1.2): {formula}")
        else:
            print(f"  Value (v1.2): {cell_v12.value}")
        
        # Compare with v1.1
        cell_v11 = sheet11[f"{col_letter}3"]
        formula_v11 = None
        if hasattr(cell_v11, 'value') and isinstance(cell_v11.value, str) and cell_v11.value.startswith('='):
            formula_v11 = cell_v11.value
        
        if formula != formula_v11:
            print(f"  [CHANGED FROM v1.1]")
            if formula_v11:
                print(f"  Formula (v1.1): {formula_v11}")
            else:
                print(f"  Value (v1.1): {cell_v11.value}")

if __name__ == '__main__':
    extract_formulas()








