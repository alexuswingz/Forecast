"""
Compare Excel versions v1.1 vs v1.2 to identify formula changes
"""
import openpyxl
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

def get_formula_from_cell(cell):
    """Extract formula from a cell, return None if no formula"""
    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
        return cell.value
    elif hasattr(cell, 'formula') and cell.formula:
        return cell.formula
    return None

def compare_excel_versions():
    """Compare v1.1 and v1.2 Excel files"""
    
    v11_path = Path("1000 Bananas AUTOFORECAST V1.1.xlsx")
    v12_path = Path("1000 Bananas AUTOFORECAST V1.2.xlsx")
    
    print("="*100)
    print("COMPARING EXCEL VERSIONS: v1.1 vs v1.2")
    print("="*100)
    
    # Load workbooks
    print("\nLoading workbooks...")
    wb11 = openpyxl.load_workbook(v11_path)
    wb12 = openpyxl.load_workbook(v12_path)
    
    print(f"\nv1.1 Sheets: {wb11.sheetnames}")
    print(f"v1.2 Sheets: {wb12.sheetnames}")
    
    # Use the "Forecast" sheet from both
    sheet11 = wb11['Forecast']
    sheet12 = wb12['Forecast']
    
    print(f"\nv1.1 Selected Sheet: {sheet11.title} ({sheet11.max_row} rows x {sheet11.max_column} cols)")
    print(f"v1.2 Selected Sheet: {sheet12.title} ({sheet12.max_row} rows x {sheet12.max_column} cols)")
    
    # Check headers first (row 1)
    print("\n" + "="*100)
    print("COLUMN HEADERS COMPARISON")
    print("="*100)
    
    headers_v11 = []
    headers_v12 = []
    
    for col in range(1, max(sheet11.max_column, sheet12.max_column) + 1):
        col_letter = get_column_letter(col)
        v11_header = sheet11[f"{col_letter}1"].value
        v12_header = sheet12[f"{col_letter}1"].value
        
        if v11_header or v12_header:
            headers_v11.append((col_letter, v11_header))
            headers_v12.append((col_letter, v12_header))
            
            if v11_header != v12_header:
                print(f"\n[CHANGED] Column {col_letter}:")
                print(f"  v1.1: {v11_header}")
                print(f"  v1.2: {v12_header}")
    
    print(f"\nTotal columns with headers:")
    print(f"  v1.1: {len([h for h in headers_v11 if h[1]])}")
    print(f"  v1.2: {len([h for h in headers_v12 if h[1]])}")
    
    # Compare formulas in each column (check row 3 onwards - row 2 is just labels)
    print("\n" + "="*100)
    print("FORMULA COMPARISON (Row 3 - First Data Row)")
    print("="*100)
    
    formula_changes = []
    
    # First, let's see what's in rows 1-3 for key columns
    print("\nFirst 3 rows structure:")
    for row in range(1, 4):
        print(f"\nRow {row}:")
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            v12_val = sheet12[f"{col_letter}{row}"].value
            v12_formula = get_formula_from_cell(sheet12[f"{col_letter}{row}"])
            if v12_formula:
                print(f"  {col_letter}: {v12_formula}")
            elif v12_val:
                print(f"  {col_letter}: {v12_val} (value)")
    
    # Now compare formulas starting from row 3
    for col in range(1, min(sheet11.max_column, sheet12.max_column, 15) + 1):  # Check first 15 cols
        col_letter = get_column_letter(col)
        
        # Get header from row 2 (seems to be the header row based on output)
        header_v11 = sheet11[f"{col_letter}2"].value
        header_v12 = sheet12[f"{col_letter}2"].value
        
        # Check row 3 (first data/formula row)
        v11_cell = sheet11[f"{col_letter}3"]
        v12_cell = sheet12[f"{col_letter}3"]
        
        v11_formula = get_formula_from_cell(v11_cell)
        v12_formula = get_formula_from_cell(v12_cell)
        
        # Only compare if at least one has a formula
        if v11_formula or v12_formula:
            if v11_formula != v12_formula:
                formula_changes.append({
                    'column': col_letter,
                    'header': header_v12 or header_v11,
                    'v11': v11_formula,
                    'v12': v12_formula
                })
                
                print(f"\n[FORMULA CHANGE] Column {col_letter}: {header_v12 or header_v11}")
                print(f"  v1.1 (Row 3): {v11_formula}")
                print(f"  v1.2 (Row 3): {v12_formula}")
    
    # Also check a middle row (row 10) to see if pattern is consistent
    print("\n" + "="*100)
    print("FORMULA VERIFICATION (Row 10 - Mid-Data Sample)")
    print("="*100)
    
    for change in formula_changes:
        col_letter = change['column']
        v11_cell = sheet11[f"{col_letter}10"]
        v12_cell = sheet12[f"{col_letter}10"]
        
        v11_formula = get_formula_from_cell(v11_cell)
        v12_formula = get_formula_from_cell(v12_cell)
        
        print(f"\n[{col_letter}] {change['header']} (Row 10):")
        print(f"  v1.1: {v11_formula}")
        print(f"  v1.2: {v12_formula}")
    
    # Check actual data values for key columns
    print("\n" + "="*100)
    print("DATA VALUES COMPARISON (First 5 rows)")
    print("="*100)
    
    # Check columns A-K for first 5 data rows
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        header = sheet12[f"{col_letter}1"].value or sheet11[f"{col_letter}1"].value
        if not header:
            continue
            
        print(f"\n[Column {col_letter}] {header}:")
        
        values_differ = False
        for row in range(2, 7):  # Rows 2-6
            v11_val = sheet11[f"{col_letter}{row}"].value
            v12_val = sheet12[f"{col_letter}{row}"].value
            
            # Check if values differ
            if v11_val != v12_val:
                values_differ = True
                print(f"  Row {row}: v1.1={v11_val} | v1.2={v12_val}")
        
        if not values_differ:
            print(f"  [No differences in first 5 rows]")
    
    # Summary
    print("\n" + "="*100)
    print("SUMMARY OF CHANGES")
    print("="*100)
    
    if formula_changes:
        print(f"\nFormula changes detected in {len(formula_changes)} columns:")
        for change in formula_changes:
            print(f"  - Column {change['column']}: {change['header']}")
    else:
        print("\n[NO FORMULA CHANGES DETECTED]")
    
    print("\n" + "="*100)
    print("CRITICAL COLUMNS TO CHECK")
    print("="*100)
    
    # Check specific important columns
    important_cols = {
        'F': 'FINAL SMOOTH',
        'G': 'FORECAST_BASELINE',
        'H': 'FORECAST_PEAK_ENV',
        'I': 'FORECAST_FINAL_SMOOTH',
        'J': 'sales_velocity_ratio',
        'K': 'sales_velocity_adj_weighted'
    }
    
    for col_letter, expected_header in important_cols.items():
        v12_header = sheet12[f"{col_letter}1"].value
        v12_formula = get_formula_from_cell(sheet12[f"{col_letter}2"])
        
        print(f"\n[{col_letter}] {v12_header}:")
        if v12_formula:
            print(f"  Formula: {v12_formula}")
        else:
            sample_val = sheet12[f"{col_letter}2"].value
            print(f"  Value: {sample_val} (no formula)")

if __name__ == '__main__':
    compare_excel_versions()

