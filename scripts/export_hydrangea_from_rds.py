import pandas as pd
from pathlib import Path
import sys

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import Config
from sqlalchemy import create_engine, text

print("=== Generating Excel from RDS ===")
print(f"RDS Host: {Config.DB_HOST}")
print(f"Database: {Config.DB_NAME}")

# Connect to RDS
rds_url = f"postgresql+psycopg2://{Config.DB_USER}:{Config.DB_PASSWORD}@{Config.DB_HOST}:{Config.DB_PORT}/{Config.DB_NAME}"
engine = create_engine(rds_url)

# Test connection
print("\nTesting RDS connection...")
try:
    with engine.connect() as conn:
        result = conn.execute(text("SELECT 1"))
        print("[OK] Connected to RDS")
except Exception as e:
    print(f"[ERROR] Could not connect to RDS: {e}")
    exit(1)

# Build the same SQL query as build_hydrangea_metrics.py
print("\nQuerying hydrangea metrics from RDS...")

query = """
WITH hyd AS (
    SELECT DISTINCT asin, UPPER(TRIM(sku)) AS sku
    FROM product_cogs
    WHERE LOWER(product_name) LIKE '%hydrangea%'
    UNION
    SELECT DISTINCT asin, UPPER(TRIM(sku)) AS sku
    FROM products
    WHERE LOWER(product_name) LIKE '%hydrangea%'
), orders AS (
    SELECT
        date_trunc('week', (order_date)::timestamptz)::date AS week_start,
        SUM(quantity) AS units,
        SUM(item_price) AS sales
    FROM order_items
    WHERE order_date IS NOT NULL
      AND (
            asin IN (SELECT asin FROM hyd WHERE asin IS NOT NULL)
            OR UPPER(sku) IN (SELECT sku FROM hyd WHERE sku IS NOT NULL)
          )
    GROUP BY 1
), traffic AS (
    SELECT
        date_trunc('week', date)::date AS week_start,
        SUM(sessions) AS sessions,
        SUM(units_ordered) AS traffic_units,
        AVG(CASE WHEN conversion_rate IS NOT NULL THEN conversion_rate ELSE NULL END) AS avg_organic_conversion_rate
    FROM child_traffic_metrics
    WHERE child_asin IN (SELECT asin FROM hyd WHERE asin IS NOT NULL)
    GROUP BY 1
), ads AS (
    SELECT
        date_trunc('week', report_date)::date AS week_start,
        SUM(impressions) AS impressions,
        SUM(clicks) AS clicks,
        SUM(spend) AS ad_spend,
        SUM(COALESCE(sales_14d, 0)) AS ad_sales,
        SUM(COALESCE(orders_14d, 0)) AS ad_orders,
        SUM(COALESCE(units_14d, 0)) AS ad_units,
        AVG(CASE WHEN conversion_rate IS NOT NULL THEN conversion_rate ELSE NULL END) AS avg_conversion_rate
    FROM ad_product_performance
    WHERE (
            advertised_asin IN (SELECT asin FROM hyd WHERE asin IS NOT NULL)
            OR UPPER(sku) IN (SELECT sku FROM hyd WHERE sku IS NOT NULL)
          )
    GROUP BY 1
), inv AS (
    SELECT
        date_trunc('week', snapshot_date)::date AS week_start,
        SUM(COALESCE(total_quantity, 0)) AS total_inventory,
        SUM(COALESCE(available_quantity, 0)) AS available_inventory,
        SUM(COALESCE(reserved_quantity, 0)) AS reserved_inventory,
        SUM(COALESCE(inbound_working_quantity, 0)) AS inbound_working,
        SUM(COALESCE(inbound_shipped_quantity, 0)) AS inbound_shipped,
        SUM(COALESCE(inbound_receiving_quantity, 0)) AS inbound_receiving,
        SUM(COALESCE(research_quantity, 0)) AS research_inventory,
        SUM(CASE WHEN fulfillment_program = 'FBA' THEN COALESCE(total_quantity, 0) ELSE 0 END) AS fba_inventory,
        SUM(CASE WHEN fulfillment_program = 'AWD' THEN COALESCE(total_quantity, 0) ELSE 0 END) AS awd_inventory
    FROM inventory_snapshots
    WHERE (
            asin IN (SELECT asin FROM hyd WHERE asin IS NOT NULL)
            OR UPPER(sku) IN (SELECT sku FROM hyd WHERE sku IS NOT NULL)
          )
    GROUP BY 1
)
SELECT
    o.week_start,
    o.units,
    o.sales,
    t.sessions,
    t.avg_organic_conversion_rate,
    a.impressions,
    a.clicks,
    a.ad_spend,
    a.ad_sales,
    a.ad_orders,
    a.ad_units,
    a.avg_conversion_rate AS ad_conversion_rate,
    i.total_inventory,
    i.available_inventory,
    i.reserved_inventory,
    i.inbound_working,
    i.inbound_shipped,
    i.inbound_receiving,
    i.research_inventory,
    i.fba_inventory,
    i.awd_inventory
FROM orders o
LEFT JOIN traffic t ON t.week_start = o.week_start
LEFT JOIN ads a ON a.week_start = o.week_start
LEFT JOIN inv i ON i.week_start = o.week_start
ORDER BY o.week_start
"""

df = pd.read_sql_query(text(query), engine)
print(f"[OK] Retrieved {len(df)} weeks of data")

# Rename columns to match expected format (PostgreSQL lowercases all unquoted identifiers)
column_mapping = {
    'week_start': 'week_start',
    'units': 'units',
    'sales': 'sales',
    'sessions': 'sessions',
    'avg_organic_conversion_rate': 'organic_conversion_rate',
    'impressions': 'ad_impressions',
    'clicks': 'ad_clicks',
    'ad_spend': 'ad_spend',
    'ad_sales': 'ad_sales',
    'ad_orders': 'ad_orders',
    'ad_units': 'ad_units',
    'ad_conversion_rate': 'ad_conversion_rate',
    'total_inventory': 'total_inventory',
    'available_inventory': 'available_inventory',
    'reserved_inventory': 'reserved_inventory',
    'inbound_working': 'inbound_working',
    'inbound_shipped': 'inbound_shipped',
    'inbound_receiving': 'inbound_receiving',
    'research_inventory': 'research_inventory',
    'fba_inventory': 'fba_inventory',
    'awd_inventory': 'awd_inventory',
}
df = df.rename(columns=column_mapping)

# Calculate derived metrics
df["ad_cpc"] = df["ad_spend"] / df["ad_clicks"]
df.loc[df["ad_clicks"] == 0, "ad_cpc"] = None
df["tacos"] = df["ad_spend"] / df["sales"]
df.loc[df["sales"] == 0, "tacos"] = None
df["organic_sales_pct"] = (df["sales"] - df["ad_sales"]) / df["sales"]
df.loc[df["sales"] == 0, "organic_sales_pct"] = None

# Create definitions dataframe
definitions_data = {
    'Metric': [
        'Week Start', 'Units', 'Sales', 'Sessions', 'Organic Conversion Rate',
        'Ad Impressions', 'Ad Clicks', 'Ad Spend', 'Ad Sales', 'Ad Orders', 'Ad Units',
        'Ad Conversion Rate', 'Total Inventory', 'Available Inventory', 'Reserved Inventory',
        'Inbound Working', 'Inbound Shipped', 'Inbound Receiving', 'Research Inventory',
        'FBA Inventory', 'AWD Inventory', 'Ad CPC', 'TACOS', 'Organic Sales %',
    ],
    'Formula/Calculation': [
        'Start date of the week (Monday)', 'Total units sold', 'Total sales revenue',
        'Total page sessions (traffic)', '(Units Ordered / Sessions) x 100',
        'Total ad impressions', 'Total ad clicks', 'Total ad spend',
        '7-day attributed ad sales', '7-day attributed ad orders', '7-day attributed ad units',
        '(Ad Orders / Ad Clicks) x 100', 'Total units across all fulfillment types',
        'Units ready to ship', 'Units reserved for orders', 'Units in inbound working status',
        'Units shipped to fulfillment centers', 'Units being received at fulfillment centers',
        'Units under research/investigation', 'Total units in FBA fulfillment centers',
        'Total units in Amazon Warehousing & Distribution', 'Ad Spend / Ad Clicks',
        '(Ad Spend / Total Sales) x 100', '((Total Sales - Ad Sales) / Total Sales) x 100',
    ],
    'Source': [
        'Calculated', 'order_items table', 'order_items table', 'child_traffic_metrics table',
        'child_traffic_metrics table', 'ad_product_performance table', 'ad_product_performance table',
        'ad_product_performance table', 'ad_product_performance table', 'ad_product_performance table',
        'ad_product_performance table', 'ad_product_performance table', 'inventory_snapshots table',
        'inventory_snapshots table', 'inventory_snapshots table', 'inventory_snapshots table',
        'inventory_snapshots table', 'inventory_snapshots table', 'inventory_snapshots table',
        'inventory_snapshots table (filtered)', 'inventory_snapshots table (filtered)',
        'Calculated', 'Calculated', 'Calculated',
    ],
    'Format': [
        'Date', 'Integer', 'Currency ($)', 'Integer', 'Percentage (%)',
        'Integer', 'Integer', 'Currency ($)', 'Currency ($)', 'Integer', 'Integer',
        'Percentage (%)', 'Integer', 'Integer', 'Integer', 'Integer',
        'Integer', 'Integer', 'Integer', 'Integer', 'Integer',
        'Currency ($)', 'Percentage (%)', 'Percentage (%)',
    ]
}
definitions_df = pd.DataFrame(definitions_data)

# Create Excel with formatting
excel_path = Path("Hydrangea_Weekly_Metrics_from_RDS.xlsx")
print(f"\nCreating Excel file: {excel_path}")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Write definitions sheet first
    definitions_df.to_excel(writer, sheet_name='Definitions', index=False)
    
    # Write weekly metrics sheet
    df.to_excel(writer, sheet_name='Weekly Metrics', index=False)
    
    # Get the workbook and worksheets
    workbook = writer.book
    definitions_ws = writer.sheets['Definitions']
    worksheet = writer.sheets['Weekly Metrics']
    
    # Format headers (bold)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Format Definitions sheet headers
    for cell in definitions_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format Weekly Metrics sheet headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format number columns
    from openpyxl.styles import numbers
    
    # Define column formats
    currency_cols = ['sales', 'ad_spend', 'ad_sales']
    percent_cols = ['organic_conversion_rate', 'ad_conversion_rate', 'tacos', 'organic_sales_pct']
    decimal_cols = ['ad_cpc']
    integer_cols = [
        'units', 'sessions', 'ad_impressions', 'ad_clicks', 'ad_orders', 'ad_units',
        'total_inventory', 'available_inventory', 'reserved_inventory',
        'inbound_working', 'inbound_shipped', 'inbound_receiving',
        'research_inventory', 'fba_inventory', 'awd_inventory'
    ]
    
    # Apply formats
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = worksheet.cell(1, col_idx).column_letter
        
        if col_name in currency_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '"$"#,##0.00'
        elif col_name in percent_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '0.00"%"'
        elif col_name in decimal_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '0.00'
        elif col_name in integer_cols:
            for row in range(2, len(df) + 2):
                worksheet[f"{col_letter}{row}"].number_format = '#,##0'
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze first row
    worksheet.freeze_panes = 'A2'
    
    # Format Definitions sheet - adjust column widths
    definitions_ws.column_dimensions['A'].width = 25  # Metric
    definitions_ws.column_dimensions['B'].width = 50  # Formula/Calculation
    definitions_ws.column_dimensions['C'].width = 35  # Source
    definitions_ws.column_dimensions['D'].width = 15  # Format
    
    # Wrap text in definitions sheet
    for row in definitions_ws.iter_rows(min_row=2, max_row=len(definitions_df) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Freeze first row in definitions sheet
    definitions_ws.freeze_panes = 'A2'

print(f"\n[SUCCESS] Excel file created: {excel_path}")
print(f"\nData sourced from RDS:")
print(f"  - {len(df)} weeks of metrics")
print(f"  - 24 columns of data")
print(f"  - 2 sheets: Definitions + Weekly Metrics")
print(f"\nFile ready: {excel_path.absolute()}")

